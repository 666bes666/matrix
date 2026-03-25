import io
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.dependencies import get_current_user, require_roles
from app.models.assessment import AggregatedScore, AssessmentCampaign
from app.models.competency import Competency
from app.models.user import User

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/matrix.xlsx")
async def export_matrix_xlsx(
    department_id: uuid.UUID | None = Query(None),
    team_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_roles("admin", "head", "department_head", "hr")),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    users_q = (
        select(User)
        .options(selectinload(User.department), selectinload(User.team))
        .where(User.is_active.is_(True))
    )
    if department_id:
        users_q = users_q.where(User.department_id == department_id)
    if team_id:
        users_q = users_q.where(User.team_id == team_id)
    users_r = await db.execute(users_q)
    users = list(users_r.scalars().all())

    comp_r = await db.execute(
        select(Competency).where(Competency.is_archived.is_(False)).order_by(Competency.name)
    )
    competencies = list(comp_r.scalars().all())

    user_ids = [u.id for u in users]
    comp_ids = [c.id for c in competencies]
    scores: dict[uuid.UUID, dict[uuid.UUID, float]] = {}
    if user_ids and comp_ids:
        agg_r = await db.execute(
            select(AggregatedScore).where(
                AggregatedScore.user_id.in_(user_ids),
                AggregatedScore.competency_id.in_(comp_ids),
            )
        )
        for agg in agg_r.scalars().all():
            scores.setdefault(agg.user_id, {})[agg.competency_id] = float(agg.final_score)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Матрица компетенций"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(1, 1, "Сотрудник").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = center
    ws.column_dimensions["A"].width = 25

    level_fills = {
        0: "C92A2A",
        1: "E67700",
        2: "FAB005",
        3: "2F9E44",
        4: "1971C2",
    }

    for col_idx, comp in enumerate(competencies, start=2):
        cell = ws.cell(1, col_idx, comp.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True, text_rotation=90
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = 6

    ws.row_dimensions[1].height = 100

    for row_idx, user in enumerate(users, start=2):
        name_cell = ws.cell(row_idx, 1, f"{user.last_name} {user.first_name}")
        name_cell.alignment = Alignment(vertical="center")
        for col_idx, comp in enumerate(competencies, start=2):
            score = scores.get(user.id, {}).get(comp.id)
            if score is not None:
                level = min(4, max(0, round(score)))
                cell = ws.cell(row_idx, col_idx, round(score, 1))
                cell.fill = PatternFill(fill_type="solid", fgColor=level_fills[level])
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=matrix.xlsx"},
    )


@router.get("/users/{user_id}/report.xlsx")
async def export_user_report_xlsx(
    user_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font

    elevated = ("admin", "head", "department_head", "hr")
    if current_user.role.value not in elevated and current_user.id != user_id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Недостаточно прав")

    user_r = await db.execute(
        select(User)
        .options(selectinload(User.department), selectinload(User.team))
        .where(User.id == user_id)
    )
    user = user_r.scalar_one_or_none()
    if user is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Пользователь не найден")

    history_r = await db.execute(
        select(AggregatedScore, AssessmentCampaign)
        .join(AssessmentCampaign, AggregatedScore.campaign_id == AssessmentCampaign.id)
        .where(AggregatedScore.user_id == user_id)
        .order_by(AssessmentCampaign.end_date.desc())
    )
    history = history_r.all()

    comp_ids = {row[0].competency_id for row in history}
    comp_names: dict[uuid.UUID, str] = {}
    if comp_ids:
        comp_r = await db.execute(select(Competency).where(Competency.id.in_(comp_ids)))
        for c in comp_r.scalars().all():
            comp_names[c.id] = c.name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт сотрудника"

    bold = Font(bold=True)
    ws.cell(1, 1, "Сотрудник:").font = bold
    ws.cell(1, 2, f"{user.last_name} {user.first_name}")
    ws.cell(2, 1, "Отдел:").font = bold
    ws.cell(2, 2, user.department.name if user.department else "—")

    ws.cell(4, 1, "Компетенция").font = bold
    ws.cell(4, 2, "Кампания").font = bold
    ws.cell(4, 3, "Итог").font = bold
    ws.cell(4, 4, "Self").font = bold
    ws.cell(4, 5, "TL").font = bold
    ws.cell(4, 6, "DH").font = bold
    ws.cell(4, 7, "Peer").font = bold
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    for row_idx, (agg, campaign) in enumerate(history, start=5):
        ws.cell(row_idx, 1, comp_names.get(agg.competency_id, str(agg.competency_id)))
        ws.cell(row_idx, 2, campaign.name)
        ws.cell(row_idx, 3, round(float(agg.final_score), 2))
        ws.cell(row_idx, 4, round(float(agg.self_score), 2) if agg.self_score is not None else "—")
        ws.cell(row_idx, 5, round(float(agg.tl_score), 2) if agg.tl_score is not None else "—")
        ws.cell(row_idx, 6, round(float(agg.dh_score), 2) if agg.dh_score is not None else "—")
        ws.cell(row_idx, 7, round(float(agg.peer_score), 2) if agg.peer_score is not None else "—")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=user_report_{user_id}.xlsx"},
    )
